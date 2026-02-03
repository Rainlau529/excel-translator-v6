# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_file, Response
import requests, time, os, tempfile, shutil, threading, uuid, json
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024

TASKS = {}
TASKS_LOCK = threading.Lock()
TASK_RETENTION_SECONDS = 1800

# ====== 限速+重试翻译 ======
def translate_text(text: str) -> str:
    if not text or not text.strip():
        return ""
    text = text[:5000]
    url = "https://translate.googleapis.com/translate_a/single"
    params = dict(client="gtx", sl="auto", tl="zh", dt="t", q=text)
    for attempt in range(1, 4):
        try:
            r = requests.get(url, params=params, timeout=15)
            if r.status_code == 200:
                data = r.json()
                return "".join(seg[0] for seg in data[0]) if data and data[0] else "翻译失败"
            time.sleep((2 ** attempt) + 0.5)
        except Exception as e:
            if attempt == 3:
                print("[ERROR] translate_text 最终失败:", e)
                return "翻译失败"
            time.sleep((2 ** attempt) + 0.5)
    return "翻译失败"

# ====== 任务/状态工具 ======
def _safe_update(task_id: str, kv: dict):
    with TASKS_LOCK:
        st = TASKS.get(task_id)
        if st:
            st.update(kv)

def _get_state(task_id: str):
    with TASKS_LOCK:
        return dict(TASKS.get(task_id) or {})

def _schedule_state_cleanup(task_id: str):
    def _cleanup():
        time.sleep(TASK_RETENTION_SECONDS)
        with TASKS_LOCK:
            TASKS.pop(task_id, None)
    threading.Thread(target=_cleanup, daemon=True).start()

# ====== 核心：8 秒心跳 + 10 行 batch ======
def _run_task(file_path: str, task_id: str):
    _safe_update(task_id, {"status": "running", "message": "读取文件...", "percent": 0, "started_at": time.time()})
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        if not ws:
            raise ValueError("空工作表")

        # 1. 找 Title 列
        title_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if str(cell.value).strip().lower() == "title":
                title_col_idx = idx
                break
        if not title_col_idx:
            raise ValueError("找不到 Title 列")

        # 2. 插入中文列
        insert_idx = title_col_idx + 1
        if ws.cell(row=1, column=insert_idx).value != "中文":
            ws.insert_cols(insert_idx)
            ws.cell(row=1, column=insert_idx, value="中文")

        # 3. 待翻译行
        rows_to_tr = [r for r in range(2, ws.max_row + 1)
                      if str(ws.cell(row=r, column=title_col_idx).value).strip()]
        total = len(rows_to_tr)
        _safe_update(task_id, {"total": total, "current": 0})
        if total == 0:
            _safe_update(task_id, {"status": "done", "percent": 100, "message": "无需翻译"})
            return

        # 4. 分页参数（8 秒必有心跳）
        BATCH = 10
        start = time.time()
        output_path = os.path.join("/tmp", os.path.splitext(os.path.basename(file_path))[0] + "_中文翻译.xlsx")

        for batch_start in range(0, total, BATCH):
            if _get_state(task_id).get("cancel_requested"):
                _safe_update(task_id, {"status": "canceled", "message": "已取消"})
                return

            batch_rows = rows_to_tr[batch_start:batch_start + BATCH]
            for done_in_batch, row in enumerate(batch_rows, 1):
                title_cell = ws.cell(row=row, column=title_col_idx)
                translated = translate_text(str(title_cell.value))
                ws.cell(row=row, column=insert_idx, value=translated)

                done_total = batch_start + done_in_batch
                elapsed = max(time.time() - start, 1e-6)
                eta = int((total - done_total) / (done_total / elapsed)) if done_total else 0
                percent = int(done_total * 100 / total)
                _safe_update(task_id, {
                    "current": done_total,
                    "percent": percent,
                    "eta_seconds": eta,
                    "message": f"翻译中({done_total}/{total})"
                })

            # 每批后保存+限速（8 秒内必有进度）
            wb.save(output_path)
            time.sleep(2)

        _safe_update(task_id, {
            "status": "done",
            "percent": 100,
            "eta_seconds": 0,
            "message": f"翻译完成，共处理 {total} 行",
            "download_filename": os.path.basename(output_path),
            "finished_at": time.time()
        })
    except Exception as e:
        _safe_update(task_id, {"status": "error", "message": str(e), "finished_at": time.time()})
    finally:
        try:
            base = os.path.dirname(file_path)
            if os.path.isdir(base):
                shutil.rmtree(base, ignore_errors=True)
        except Exception:
            pass
        _schedule_state_cleanup(task_id)

# ====== 其余路由/启动代码 ======
def start_background_task(file_path: str, filename: str) -> str:
    task_id = uuid.uuid4().hex[:12]
    with TASKS_LOCK:
        TASKS[task_id] = {
            "status": "idle", "percent": 0, "total": 0, "current": 0,
            "cancel_requested": False, "tmp_dir": os.path.dirname(file_path),
            "filename": filename
        }
    threading.Thread(target=_run_task, args=(file_path, task_id), daemon=True).start()
    return task_id

def sse_progress(task_id: str):
    def gen():
        while True:
            st = _get_state(task_id)
            if not st:
                yield f"data: {json.dumps({'status':'error','message':'任务不存在'})}\n\n"
                break
            payload = {
                "task_id": task_id,
                "status": st.get("status"),
                "percent": st.get("percent"),
                "eta_seconds": st.get("eta_seconds"),
                "message": st.get("message"),
                "download_url": f"/download/{st.get('download_filename')}" if st.get("download_filename") else None,
                "filename": st.get("filename"),
                "started_at": st.get("started_at"),
                "finished_at": st.get("finished_at"),
                "duration_seconds": st.get("duration_seconds")
            }
            yield f"event: progress\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"
            if st.get("status") in ("done", "error", "canceled"):
                break
            time.sleep(0.2)
    return Response(gen(), mimetype="text/event-stream")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files.get('file')
    if not file or file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传Excel文件(.xlsx)'}), 400
    try:
        filename = secure_filename(file.filename)
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, filename)
        file.save(file_path)
        task_id = start_background_task(file_path, filename)
        return jsonify({'success': True, 'task_id': task_id}), 202
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/progress/<task_id>')
def progress(task_id):
    return sse_progress(task_id)

@app.route('/tasks/<task_id>', methods=['GET'])
def task_status(task_id):
    st = _get_state(task_id)
    return jsonify(st) if st else (jsonify({'error': '任务不存在'}), 404)

@app.route('/download/<filename>')
def download_file(filename):
    # 文件持久化在 /tmp
    file_path = os.path.join("/tmp", filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    return "文件不存在", 404
